"""
행사 데이터 수집용 엑셀 템플릿 생성 스크립트
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 스타일 정의
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

def apply_section_style(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

def apply_cell_style(cell):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')

def create_promotion_master_sheet(wb):
    """행사 마스터 시트 생성"""
    ws = wb.create_sheet("행사정보")

    # 헤더 정의
    headers = [
        ("A", "행사명*", 25),
        ("B", "행사유형*", 25),
        ("C", "시작일*\n(YYYY-MM-DD)", 15),
        ("D", "종료일*\n(YYYY-MM-DD)", 15),
        ("E", "상태", 12),
        ("F", "브랜드ID*", 12),
        ("G", "채널명*", 20),
        ("H", "채널수수료율\n(%)", 15),
        ("I", "할인분담주체", 15),
        ("J", "오리오분담률\n(%)", 15),
        ("K", "채널분담률\n(%)", 15),
        ("L", "목표매출", 15),
        ("M", "목표수량", 12),
        ("N", "목표순이익", 15),
        ("O", "비고", 30),
    ]

    # 헤더 적용
    for col, (letter, header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        set_column_width(ws, col, width)

    # 행사유형 드롭다운
    promotion_types = [
        "ONLINE_PRICE_DISCOUNT",
        "ONLINE_COUPON",
        "ONLINE_POST_SETTLEMENT",
        "ONLINE_FREE_PROMOTION",
        "OFFLINE_WHOLESALE_DISCOUNT",
        "OFFLINE_SPECIAL_PRODUCT",
        "OFFLINE_BUNDLE_DISCOUNT"
    ]
    dv_type = DataValidation(
        type="list",
        formula1=f'"{",".join(promotion_types)}"',
        allow_blank=True
    )
    dv_type.error = "목록에서 선택해주세요"
    dv_type.errorTitle = "잘못된 입력"
    ws.add_data_validation(dv_type)
    dv_type.add("B2:B100")

    # 상태 드롭다운
    status_list = ["SCHEDULED", "ACTIVE", "COMPLETED", "CANCELLED"]
    dv_status = DataValidation(
        type="list",
        formula1=f'"{",".join(status_list)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("E2:E100")

    # 할인분담주체 드롭다운
    owner_list = ["COMPANY", "CHANNEL", "BOTH"]
    dv_owner = DataValidation(
        type="list",
        formula1=f'"{",".join(owner_list)}"',
        allow_blank=True
    )
    ws.add_data_validation(dv_owner)
    dv_owner.add("I2:I100")

    # 데이터 영역 스타일
    for row in range(2, 11):
        for col in range(1, len(headers) + 1):
            apply_cell_style(ws.cell(row=row, column=col))

    # 행 높이
    ws.row_dimensions[1].height = 35

    return ws

def create_product_sheet(wb):
    """행사 상품 시트 생성"""
    ws = wb.create_sheet("행사상품")

    # 헤더 정의 (섹션별로 구분)
    headers = [
        # 기본 정보
        ("행사명*", 20),
        ("상품ID*", 12),
        ("상품명\n(참고용)", 25),
        # 가격 구조
        ("정가", 12),
        ("상시판매가", 12),
        ("행사가\n(판매가할인)", 15),
        ("공급가\n(원매가할인)", 15),
        ("쿠폰할인율\n(%)", 12),
        # 비용 구조
        ("상품원가", 12),
        ("정산방식", 15),
        ("물류비", 12),
        ("관리비", 12),
        ("창고비", 12),
        ("EDI비용", 12),
        # 목표
        ("목표매출", 15),
        ("목표수량", 12),
        ("필요재고", 12),
        # 기타
        ("비고", 30),
    ]

    # 섹션 헤더 (2행에 병합 셀로 표시)
    sections = [
        (1, 3, "기본 정보"),
        (4, 8, "가격 구조"),
        (9, 14, "비용 구조"),
        (15, 17, "목표"),
        (18, 18, "기타"),
    ]

    # 섹션 헤더 적용 (1행)
    for start, end, name in sections:
        cell = ws.cell(row=1, column=start, value=name)
        apply_section_style(cell)
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for c in range(start, end + 1):
                apply_section_style(ws.cell(row=1, column=c))

    # 컬럼 헤더 적용 (2행)
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        apply_header_style(cell)
        set_column_width(ws, col, width)

    # 정산방식 드롭다운
    settlement_list = ["DIRECT", "CONSIGNMENT"]
    dv_settlement = DataValidation(
        type="list",
        formula1=f'"{",".join(settlement_list)}"',
        allow_blank=True
    )
    dv_settlement.error = "DIRECT 또는 CONSIGNMENT를 선택해주세요"
    ws.add_data_validation(dv_settlement)
    dv_settlement.add("J3:J200")

    # 데이터 영역 스타일
    for row in range(3, 21):
        for col in range(1, len(headers) + 1):
            apply_cell_style(ws.cell(row=row, column=col))

    # 행 높이
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 35

    return ws

def create_guide_sheet(wb):
    """입력 가이드 시트 생성"""
    ws = wb.create_sheet("입력가이드")

    guide_content = [
        ("행사 데이터 입력 가이드", None),
        ("", None),
        ("1. 행사유형 설명", None),
        ("", None),
        ("온라인", None),
        ("유형코드", "설명"),
        ("ONLINE_PRICE_DISCOUNT", "판매가 할인 (상시판매가 → 행사가로 할인)"),
        ("ONLINE_COUPON", "쿠폰 할인 (쿠폰할인율 적용)"),
        ("ONLINE_POST_SETTLEMENT", "정산후보정 (후정산 방식)"),
        ("ONLINE_FREE_PROMOTION", "무상기획전"),
        ("", None),
        ("오프라인", None),
        ("유형코드", "설명"),
        ("OFFLINE_WHOLESALE_DISCOUNT", "원매가 할인 (공급가 할인)"),
        ("OFFLINE_SPECIAL_PRODUCT", "기획상품 (특별 기획 상품)"),
        ("OFFLINE_BUNDLE_DISCOUNT", "에누리/묶음할인 (예: 3개 구매시 20% 할인)"),
        ("", None),
        ("2. 필수 입력 필드 (*표시)", None),
        ("", None),
        ("행사정보 시트:", "행사명, 행사유형, 시작일, 종료일, 브랜드ID, 채널명"),
        ("행사상품 시트:", "행사명, 상품ID"),
        ("", None),
        ("3. 유형별 필수 가격 필드", None),
        ("", None),
        ("ONLINE_PRICE_DISCOUNT", "상시판매가, 행사가(판매가할인)"),
        ("ONLINE_COUPON", "상시판매가, 쿠폰할인율"),
        ("OFFLINE_WHOLESALE_DISCOUNT", "상시판매가, 공급가(원매가할인)"),
        ("OFFLINE_SPECIAL_PRODUCT", "행사가(기획상품가)"),
        ("OFFLINE_BUNDLE_DISCOUNT", "상시판매가, 행사가(묶음할인가)"),
        ("", None),
        ("4. 정산방식", None),
        ("", None),
        ("DIRECT", "완사입 (직접 매입)"),
        ("CONSIGNMENT", "위탁 (수수료 방식)"),
        ("", None),
        ("5. 할인분담주체", None),
        ("", None),
        ("COMPANY", "오리오 전액 부담"),
        ("CHANNEL", "채널 전액 부담"),
        ("BOTH", "오리오/채널 분담 (분담률 입력 필요)"),
    ]

    # 제목 스타일
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)

    for row_idx, (col1, col2) in enumerate(guide_content, 1):
        cell1 = ws.cell(row=row_idx, column=1, value=col1)
        if col2:
            ws.cell(row=row_idx, column=2, value=col2)

        # 스타일 적용
        if row_idx == 1:
            cell1.font = title_font
        elif col1 and (col1.startswith(("1.", "2.", "3.", "4.", "5.")) or col1 in ["온라인", "오프라인"]):
            cell1.font = section_font

    set_column_width(ws, 1, 30)
    set_column_width(ws, 2, 50)

    return ws

def create_example_sheet(wb):
    """예시 데이터 시트 생성"""
    ws = wb.create_sheet("예시데이터")

    # 행사 예시
    ws.cell(row=1, column=1, value="[행사정보 예시]").font = Font(bold=True, size=12)

    promo_headers = ["행사명", "행사유형", "시작일", "종료일", "브랜드ID", "채널명", "채널수수료율", "할인분담주체", "오리오분담률"]
    promo_examples = [
        ["쿠팡 로켓배송 특가", "ONLINE_PRICE_DISCOUNT", "2026-02-01", "2026-02-07", 1, "쿠팡", 15.0, "BOTH", 60.0],
        ["네이버 쿠폰 행사", "ONLINE_COUPON", "2026-02-01", "2026-02-14", 1, "네이버스토어", 12.0, "COMPANY", 100.0],
        ["롯데마트 원매가 할인", "OFFLINE_WHOLESALE_DISCOUNT", "2026-02-15", "2026-02-21", 1, "롯데마트", None, "COMPANY", 100.0],
        ["이마트 3개 구매 할인", "OFFLINE_BUNDLE_DISCOUNT", "2026-02-01", "2026-02-28", 1, "이마트, 홈플러스", None, "COMPANY", 100.0],
    ]

    for col, header in enumerate(promo_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        apply_header_style(cell)
        set_column_width(ws, col, 20)

    for row_idx, example in enumerate(promo_examples, 3):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    # 상품 예시
    ws.cell(row=9, column=1, value="[행사상품 예시]").font = Font(bold=True, size=12)

    prod_headers = ["행사명", "상품ID", "상시판매가", "행사가", "공급가", "쿠폰할인율", "상품원가", "정산방식", "물류비"]
    prod_examples = [
        ["쿠팡 로켓배송 특가", 101, 10000, 8500, None, None, 5000, "CONSIGNMENT", 500],
        ["네이버 쿠폰 행사", 101, 10000, None, None, 15.0, 5000, "CONSIGNMENT", 500],
        ["롯데마트 원매가 할인", 101, 10000, 10000, 6000, None, 5000, "DIRECT", 300],
        ["이마트 3개 구매 할인", 101, 10000, 8000, 7000, None, 5000, "DIRECT", 300],
    ]

    for col, header in enumerate(prod_headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        apply_header_style(cell)

    for row_idx, example in enumerate(prod_examples, 11):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)

    return ws

def main():
    wb = Workbook()

    # 기본 시트 삭제
    wb.remove(wb.active)

    # 시트 생성
    create_guide_sheet(wb)
    create_promotion_master_sheet(wb)
    create_product_sheet(wb)
    create_example_sheet(wb)

    # 저장
    output_path = r"C:\Python\Azure\Future_Mockups\promotion_data_template.xlsx"
    wb.save(output_path)
    print(f"엑셀 템플릿 생성 완료: {output_path}")

if __name__ == "__main__":
    main()