"""
RevenuePlan 엑셀 업로드 스크립트
- 엑셀 파일(Date, BrandName, ChannelName, ChannelDetail, PlanType, Amount)을 읽어
  RevenuePlan 테이블에 업로드
- 동일 Date + BrandID + ChannelID + PlanType + ChannelDetail 조합이 이미 있으면 UPDATE, 없으면 INSERT
"""

import sys
import openpyxl
from WebApp_v2_admin.core.database import get_db_cursor


def load_excel(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = ['Date', 'BrandName', 'ChannelName', 'ChannelDetail', 'PlanType', 'Amount']
    if headers != expected:
        print(f"[ERROR] 헤더 불일치. 예상: {expected}, 실제: {headers}")
        sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({
            'Date': row[0],
            'BrandName': row[1],
            'ChannelName': row[2],
            'ChannelDetail': row[3],
            'PlanType': row[4],
            'Amount': row[5],
        })
    return rows


def build_lookup(cursor) -> tuple[dict, dict]:
    """Brand, Channel 이름 -> ID 매핑 딕셔너리 생성"""
    cursor.execute("SELECT BrandID, Name FROM Brand")
    brand_map = {name.strip(): bid for bid, name in cursor.fetchall()}

    cursor.execute("SELECT ChannelID, Name FROM Channel")
    channel_map = {name.strip(): cid for cid, name in cursor.fetchall()}

    return brand_map, channel_map


def validate(rows: list[dict], brand_map: dict, channel_map: dict) -> list[str]:
    """전체 데이터 사전 검증. 오류 목록 반환 (빈 리스트면 통과)"""
    errors = []
    for i, r in enumerate(rows, start=2):
        brand_name = str(r['BrandName']).strip()
        channel_name = str(r['ChannelName']).strip()

        if brand_name not in brand_map:
            errors.append(f"  [ROW {i}] 브랜드 '{brand_name}' 없음")
        if channel_name not in channel_map:
            errors.append(f"  [ROW {i}] 채널 '{channel_name}' 없음")

        plan_type = str(r['PlanType']).strip().upper()
        if plan_type not in ('EXPECTED', 'TARGET'):
            errors.append(f"  [ROW {i}] PlanType '{r['PlanType']}' 잘못됨 (EXPECTED/TARGET만 가능)")

        if r['Date'] is None:
            errors.append(f"  [ROW {i}] Date 값 없음")

    return errors


def upload(file_path: str):
    rows = load_excel(file_path)
    if not rows:
        print("[INFO] 업로드할 데이터가 없습니다.")
        return

    print(f"[INFO] {len(rows)}건 로드 완료")

    with get_db_cursor(commit=True) as cursor:
        brand_map, channel_map = build_lookup(cursor)

        # 1단계: 사전 검증
        errors = validate(rows, brand_map, channel_map)
        if errors:
            print(f"\n[ERROR] 검증 실패 ({len(errors)}건) - 업로드를 중단합니다.")
            print("엑셀 파일을 수정한 후 다시 실행하세요.\n")
            for e in errors:
                print(e)
            return

        print("[INFO] 검증 통과. 업로드를 시작합니다.\n")

        # 2단계: 업로드
        success = 0
        for i, r in enumerate(rows, start=2):
            brand_id = brand_map[str(r['BrandName']).strip()]
            channel_id = channel_map[str(r['ChannelName']).strip()]
            plan_type = str(r['PlanType']).strip().upper()
            amount = float(r['Amount'] or 0)
            channel_detail = r['ChannelDetail']

            cursor.execute("""
                SELECT PlanID FROM RevenuePlan
                WHERE [Date] = ? AND BrandID = ? AND ChannelID = ?
                  AND PlanType = ? AND ISNULL(ChannelDetail, '') = ISNULL(?, '')
            """, (r['Date'], brand_id, channel_id, plan_type, channel_detail))

            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE RevenuePlan
                    SET Amount = ?, UpdatedAt = GETDATE()
                    WHERE PlanID = ?
                """, (amount, existing[0]))
                print(f"  [ROW {i}] UPDATE PlanID={existing[0]} Amount={amount}")
            else:
                cursor.execute("""
                    INSERT INTO RevenuePlan ([Date], BrandID, ChannelID, PlanType, Amount, ChannelDetail)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (r['Date'], brand_id, channel_id, plan_type, amount, channel_detail))
                print(f"  [ROW {i}] INSERT {r['Date']} {r['BrandName']}/{r['ChannelName']} {plan_type} {amount}")

            success += 1

    print(f"\n[완료] {success}건 업로드 성공")


if __name__ == '__main__':
    file_path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Python\RevenuePlan_Upload.xlsx'
    print(f"[INFO] 파일: {file_path}")
    upload(file_path)
